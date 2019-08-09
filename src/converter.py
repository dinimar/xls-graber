import csv
import datetime
import logging
import os
import configparser
import utils.grab_lib as graber
from openpyxl import Workbook

# create logger
cli_logger = logging.getLogger('i-otvet-converter')
# set logging level
cli_logger.setLevel(logging.DEBUG)
# create console and file handler
ch = logging.StreamHandler()
# create formatter
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
# add formatter to handlers
ch.setFormatter(formatter)
# add handlers to logger
cli_logger.addHandler(ch)

root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

config_file = "graber.ini"
config = configparser.ConfigParser()
config.read(os.path.join(root, config_file))
q_csv_dir = config["paths"]["q_csv_dir"]
q_csv_file_name = config["paths"]["q_csv_file_name"]
q_xlsx_dir = config["paths"]["q_xlsx_dir"]
# create xlsx dir if there is no such dir
if not os.path.exists(os.path.join(root, q_xlsx_dir)):
    os.mkdir(os.path.join(root, q_xlsx_dir))
q_xlsx_file_name = config["paths"]["q_xlsx_file_name"]


save_coef = int(config["general"]["save_coef"])

# xlsx_file = open(os.path.join(root, out_dir+xlsx_name), 'w')
# xlsx_file.close()

fieldnames = {'Id': 1, 'Type': 2, 'ParentIdInFile': 3, 'ParentIdInSIte': 4,
              'Title': 5, 'Content': 6, 'Format': 7, 'CategoryId': 8, 'CategoryUrl': 9,
              'Tags': 10, 'UserName': 11, 'AnonymousName': 12, 'Notify': 13, 'ExtraValue': 14,
              'DateTimeFrom': 15, 'DateTimeTo': 16, 'Selected': 17}
date_p = "%m/%d/%Y %H:%M:%S"


def get_date(date_str):
    # parse string
    return datetime.datetime.strptime(date_str, date_p)


def create_xls(csv_name, out_xls_name):
    last_free_row = 1
    wb_q = Workbook()

    with open(os.path.join(root, csv_name), newline='', encoding='utf-8') as csvfile:
        # reader = csv.DictReader(csvfile)
        spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
        wb_q_s = wb_q.active
        # add header for redundant columns
        # remove ParentIdInSite(redundant)
        wb_q_s.cell(column=fieldnames['ParentIdInSIte'], row=last_free_row).value = "ParentIdInSIte"
        # remove CategoryUrl, Tags, UserName(redundant)
        wb_q_s.cell(column=fieldnames['CategoryUrl'], row=last_free_row).value = "CategoryUrl"
        wb_q_s.cell(column=fieldnames['Tags'], row=last_free_row).value = "Tags"
        wb_q_s.cell(column=fieldnames['UserName'], row=last_free_row).value = "UserName"
        # remove Notify, ExtraValue(redundant)
        wb_q_s.cell(column=fieldnames['Notify'], row=last_free_row).value = "Notify"
        wb_q_s.cell(column=fieldnames['ExtraValue'], row=last_free_row).value = "ExtraValue"
        # remove DateTimeTo, Selected(redundant)
        wb_q_s.cell(column=fieldnames['DateTimeTo'], row=last_free_row).value = "DateTimeTo"
        wb_q_s.cell(column=fieldnames['Selected'], row=last_free_row).value = "Selected"
        for row in spamreader:
            # if it's not empty and a first row parse to int
            if (row[fieldnames['Id'] - 1] != "" and row[fieldnames['Id'] - 1] != "Id"):
                wb_q_s.cell(column=fieldnames['Id'], row=last_free_row).value = int(row[fieldnames['Id'] - 1])
            elif (row[fieldnames['Id'] - 1] == "Id"):
                wb_q_s.cell(column=fieldnames['Id'], row=last_free_row).value = row[fieldnames['Id'] - 1]
            wb_q_s.cell(column=fieldnames['Type'], row=last_free_row).value = row[fieldnames['Type'] - 1]
            # add ParentIdInFile only for answers and first row
            if (row[fieldnames['ParentIdInFile'] - 1] != ""):
                wb_q_s.cell(column=fieldnames['ParentIdInFile'], row=last_free_row).value = row[
                    fieldnames['ParentIdInFile'] - 1]
            # set title if it's a question
            if (row[fieldnames['Type'] - 1] == "Q"):
                wb_q_s.cell(column=fieldnames['Title'], row=last_free_row).value = row[fieldnames['Title'] - 1]
            # if there is no content -> 'N/A'
            if (row[fieldnames['Content'] - 1] == ""):
                wb_q_s.cell(column=fieldnames['Content'], row=last_free_row).value = "N/A"
            else:
                wb_q_s.cell(column=fieldnames['Content'], row=last_free_row).value = row[fieldnames['Content'] - 1]
            wb_q_s.cell(column=fieldnames['Format'], row=last_free_row).value = row[fieldnames['Format'] - 1]
            # if it's a question and not a first row
            if (row[fieldnames['CategoryId'] - 1] != "" and row[fieldnames['CategoryId'] - 1] != "CategoryId"):
                wb_q_s.cell(column=fieldnames['CategoryId'], row=last_free_row).value = int(
                    row[fieldnames['CategoryId'] - 1])
            elif (row[fieldnames['CategoryId'] - 1] == 'CategoryId'):
                wb_q_s.cell(column=fieldnames['CategoryId'], row=last_free_row).value = row[
                    fieldnames['CategoryId'] - 1]
            wb_q_s.cell(column=fieldnames['AnonymousName'], row=last_free_row).value = row[
                fieldnames['AnonymousName'] - 1]
            # process date str
            if (row[fieldnames['DateTimeFrom'] - 1] != "DateTimeFrom"):
                wb_q_s.cell(column=fieldnames['DateTimeFrom'], row=last_free_row).value = get_date(
                    row[fieldnames['DateTimeFrom'] - 1])
            last_free_row = last_free_row + 1

        wb_q.save(filename=os.path.join(root, out_xls_name))

    cli_logger.info(out_xls_name + " is created")


idx = save_coef
csv_file_name = os.path.join(root, (q_csv_dir + str(idx) + '-' + q_csv_file_name))
xls_file_name = os.path.join(root, (q_xlsx_dir + str(idx) + '-' + q_xlsx_file_name))

while (os.path.isfile(csv_file_name)):
    create_xls(csv_file_name, xls_file_name)
    idx = idx + save_coef
    csv_file_name = os.path.join(root, (q_csv_dir + str(idx) + '-' + q_csv_file_name))
    xls_file_name = os.path.join(root, (q_xlsx_dir + str(idx) + '-' + q_xlsx_file_name))
# break
